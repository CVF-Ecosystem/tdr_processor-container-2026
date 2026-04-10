# dashboard_shared.py
"""Shared utilities for TDR multi-page Streamlit dashboard."""
import io
import json
import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from config import settings as _settings
    KPI_MOVES_PER_HOUR: int = _settings.kpi_moves_per_hour
    _data_dir = _settings.data_dir
    _db_path = _settings.db_path
except Exception:
    KPI_MOVES_PER_HOUR = 45
    _data_dir = "outputs/data_csv"
    _db_path = "outputs/tdr_master.db"

DATA_DIR = Path(_data_dir)
DB_PATH = Path(_db_path)

# AgGrid — graceful fallback nếu chưa cài
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode  # type: ignore[import]
    _AGGRID_AVAILABLE = True
except ImportError:
    _AGGRID_AVAILABLE = False


# ── Plotly chart theme ────────────────────────────────────────────────────────
_CHART_PALETTE = ["#2563EB", "#16A34A", "#DC2626", "#D97706", "#7C3AED",
                  "#0891B2", "#DB2777", "#65A30D", "#EA580C", "#0F172A"]


def apply_chart_theme(fig, title: str = ""):
    """S4-5: Áp dụng TDR chart theme thống nhất cho mọi Plotly figure."""
    fig.update_layout(
        template="plotly_white",
        font=dict(family="Segoe UI, Arial, sans-serif", size=12),
        title=dict(text=title, font=dict(size=14, color="#1e293b")) if title else {},
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        colorway=_CHART_PALETTE,
        xaxis=dict(showgrid=False, linecolor="#e2e8f0", tickfont=dict(size=11)),
        yaxis=dict(gridcolor="#f1f5f9", linecolor="#e2e8f0", tickfont=dict(size=11)),
        hoverlabel=dict(bgcolor="#ffffff", bordercolor="#cbd5e1", font=dict(size=12)),
        legend=dict(bgcolor="rgba(0,0,0,0)", borderwidth=0),
        margin=dict(t=50, l=50, r=20, b=50),
    )
    return fig


# ── KPI Card ─────────────────────────────────────────────────────────────────
def kpi_card(title: str, value: str, delta: str = "", color: str = "#2563EB") -> None:
    """S4-1: Render KPI card với border màu, typography rõ ràng."""
    delta_html = f'<div style="font-size:12px;color:{color};margin-top:4px">{delta}</div>' if delta else ""
    st.markdown(f"""
    <div style="border-left:4px solid {color};padding:12px 16px;
                border-radius:6px;background:#f8fafc;margin-bottom:8px;
                box-shadow:0 1px 3px rgba(0,0,0,.06)">
      <div style="font-size:11px;color:#64748b;font-weight:600;
                  text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">{title}</div>
      <div style="font-size:26px;font-weight:700;color:#1e293b;line-height:1">{value}</div>
      {delta_html}
    </div>""", unsafe_allow_html=True)


# ── Global CSS injection (S4-4 dark mode + S4-8 branding cleanup) ────────────
def inject_global_css() -> None:
    """S4-8: Ẩn Streamlit mặc định. S4-4: Dark mode toggle trong sidebar."""
    # Branding/cleanup CSS — luôn áp dụng
    base_css = """<style>
    #MainMenu{visibility:hidden}
    footer{visibility:hidden}
    [data-testid="stToolbar"]{display:none}
    .stDeployButton{display:none}
    /* Sidebar logo area */
    [data-testid="stSidebarHeader"]{padding-bottom:0}
    </style>"""
    st.markdown(base_css, unsafe_allow_html=True)

    # Dark mode toggle
    dark = st.sidebar.toggle("🌙 Dark mode", key="dark_mode", value=False)
    if dark:
        st.markdown("""<style>
        .stApp{background-color:#1e1e2e!important}
        section[data-testid="stSidebar"]>div{background-color:#252535!important}
        .stMarkdown p,.stMarkdown h1,.stMarkdown h2,.stMarkdown h3,
        label,.stTextInput label,.stSelectbox label{color:#cdd6f4!important}
        </style>""", unsafe_allow_html=True)


# ── AgGrid helper (S4-6) ──────────────────────────────────────────────────────
def render_aggrid_table(df: pd.DataFrame, height: int = 400,
                        page_size: int = 50, fit_columns: bool = True) -> None:
    """S4-6: Render bảng AgGrid; fallback về st.dataframe nếu chưa cài package."""
    if not _AGGRID_AVAILABLE:
        st.dataframe(df, use_container_width=True)
        return
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        filterable=True, sortable=True, resizable=True, min_column_width=80
    )
    gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=page_size)
    gb.configure_side_bar(filters_panel=True, columns_panel=False)
    gb.configure_selection("single")
    grid_opts = gb.build()
    AgGrid(
        df, gridOptions=grid_opts, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=fit_columns,
        use_container_width=True,
    )


# ── Translations & Language ───────────────────────────────────────────────────
@st.cache_data
def load_translations(lang_file: str = "locales.json") -> dict:
    with open(lang_file, "r", encoding="utf-8") as f:
        return json.load(f)


def t(key: str, **kwargs) -> str:
    """Return translated string for current language, fallback vi → raw key."""
    translations = load_translations()
    lang = st.session_state.get("lang", "vi")
    text = translations[lang].get(key) or translations["vi"].get(key, key)
    return text.format(**kwargs) if kwargs else text


def init_lang() -> None:
    """Initialize language session state (call at top of every page)."""
    if "lang" not in st.session_state:
        st.session_state.lang = "vi"


# ── Data Loading ──────────────────────────────────────────────────────────────
@st.cache_data(ttl=30)
def load_csv(path_str: str) -> pd.DataFrame:
    """Load CSV with 30-second TTL — auto-refreshes when new data is processed."""
    path = Path(path_str)
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception as e:
        st.error(f"Error reading {path.name}: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=30)
def load_table(table: str) -> pd.DataFrame:
    """Load a table from SQLite (30s TTL). Falls back to CSV if DB not found."""
    if not DB_PATH.exists():
        csv_fallback = {
            "vessel_summary":           "vessel_summary.csv",
            "qc_productivity":          "qc_productivity.csv",
            "qc_operator_productivity": "qc_operator_productivity.csv",
            "delay_details":            "delay_details.csv",
            "container_details_long":   "container_details_long.csv",
        }
        csv_name = csv_fallback.get(table)
        return load_csv(str(DATA_DIR / csv_name)) if csv_name else pd.DataFrame()
    _CSV_FALLBACK = {
        "vessel_summary":           "vessel_summary.csv",
        "qc_productivity":          "qc_productivity.csv",
        "qc_operator_productivity": "qc_operator_productivity.csv",
        "delay_details":            "delay_details.csv",
        "container_details_long":   "container_details_long.csv",
    }
    try:
        con = sqlite3.connect(DB_PATH)
        df = pd.read_sql(f"SELECT * FROM {table}", con)  # noqa: S608
        con.close()
        return df
    except Exception:
        csv_name = _CSV_FALLBACK.get(table)
        if csv_name:
            df_csv = load_csv(str(DATA_DIR / csv_name))
            if not df_csv.empty:
                return df_csv
        st.warning(f"[SQLite] Không tìm thấy dữ liệu '{table}'. Chạy xử lý TDR để tạo dữ liệu.")
        return pd.DataFrame()


_VESSEL_NUMERIC_COLS = [
    'Grand Total Conts', 'Grand Total TEUs', 'Portstay (hrs)',
    'Gross Working (hrs)', 'Net Working (hrs)', 'Break time (hrs)',
    'Discharge (hrs)', 'Loading (hrs)', 'Total working (hrs)',
    'Vessel Moves/Portstay Hour', 'Vessel Moves/Gross Hour', 'Vessel Moves/Net Hour',
]


def _coerce_vessel_types(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure numeric columns are float/int — guards against object dtype from SQLite/CSV."""
    for col in _VESSEL_NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def load_all_data():
    """Return (df_vessel_raw, df_qc, df_delay, df_container, df_qc_operator).
    Ưu tiên SQLite; fallback CSV khi DB chưa có."""
    df_vessel = _coerce_vessel_types(load_table("vessel_summary"))
    return (
        df_vessel,
        load_table("qc_productivity"),
        load_table("delay_details"),
        load_table("container_details_long"),
        load_table("qc_operator_productivity"),
    )


# ── Sidebar ───────────────────────────────────────────────────────────────────
def render_language_selector() -> None:
    """Render language radio + CSS injection (dark mode, branding cleanup)."""
    st.sidebar.title("⚙️ Settings")
    inject_global_css()
    lang_options = {"Tiếng Việt": "vi", "English": "en"}
    current_key = next(
        (k for k, v in lang_options.items() if v == st.session_state.get("lang", "vi")),
        "Tiếng Việt",
    )
    selected = st.sidebar.radio(
        "Ngôn ngữ / Language",
        options=list(lang_options.keys()),
        index=list(lang_options.keys()).index(current_key),
        key="lang_radio",
    )
    if st.session_state.lang != lang_options[selected]:
        st.session_state.lang = lang_options[selected]
        st.rerun()


def render_sidebar_filters(df_vessel_raw: pd.DataFrame) -> pd.DataFrame:
    """Render filter widgets and return filtered vessel DataFrame."""
    st.sidebar.markdown("---")
    st.sidebar.subheader(t("filter_header"))

    df = df_vessel_raw.copy()
    if "Report Date" in df.columns:
        df["Report Date"] = pd.to_datetime(df["Report Date"], errors="coerce")
        df["YearMonth"] = df["Report Date"].dt.to_period("M").astype(str)
        df["Quarter"] = df["Report Date"].dt.to_period("Q").astype(str)

    sel_op = t("filter_all")
    if "Operator" in df.columns:
        ops = [t("filter_all")] + sorted(df["Operator"].dropna().unique().tolist())
        sel_op = st.sidebar.selectbox(t("filter_operator"), ops, key="sb_operator")

    sel_berth = t("filter_all")
    if "Berth" in df.columns:
        berths = [t("filter_all")] + sorted(df["Berth"].dropna().unique().tolist())
        sel_berth = st.sidebar.selectbox(t("filter_berth"), berths, key="sb_berth")

    date_range = None
    if "Report Date" in df.columns and df["Report Date"].notna().any():
        min_d = df["Report Date"].min().date()
        max_d = df["Report Date"].max().date()
        date_range = st.sidebar.date_input(
            t("filter_date_range"),
            value=(min_d, max_d),
            min_value=min_d,
            max_value=max_d,
            key="sb_date_range",
        )

    if sel_op != t("filter_all"):
        df = df[df["Operator"] == sel_op]
    if sel_berth != t("filter_all"):
        df = df[df["Berth"] == sel_berth]
    if date_range and len(date_range) == 2:
        s = pd.Timestamp(date_range[0])
        e = pd.Timestamp(date_range[1]) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        df = df[(df["Report Date"] >= s) & (df["Report Date"] <= e)]

    if len(df) < len(df_vessel_raw):
        st.sidebar.info(f"📊 {t('filtered_vessels')}: {len(df)}/{len(df_vessel_raw)}")

    if st.sidebar.button(t("filter_reset"), key="btn_reset"):
        for key in ("sb_operator", "sb_berth", "sb_date_range"):
            st.session_state.pop(key, None)
        st.rerun()

    return df


# ── Export ────────────────────────────────────────────────────────────────────
def _create_styled_excel(tables: list) -> io.BytesIO:
    """S4-2: Excel export với header formatting, auto column width, conditional formatting."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in tables:
            if df.empty:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]

            # Header row: bold + nền xanh navy + chữ trắng
            hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
            hdr_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align

            # Auto column width (max 45)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

            # Freeze top row
            ws.freeze_panes = "A2"

            # Conditional formatting: đỏ nhạt nếu Net moves/h < KPI
            col_names = [c.value for c in ws[1]]
            if "Net moves/h" in col_names:
                col_idx = col_names.index("Net moves/h") + 1
                col_letter = ws.cell(1, col_idx).column_letter
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator="lessThan", formula=[str(KPI_MOVES_PER_HOUR)], fill=red_fill),
                )

    buf.seek(0)
    return buf


def render_export_sidebar(df_vessel_raw, df_qc, df_qc_operator, df_delay, df_container) -> None:
    """Render export section at bottom of sidebar."""
    st.sidebar.markdown("---")
    st.sidebar.subheader(t("export_header"))

    export_fmt = st.sidebar.selectbox(t("export_format"), ["CSV", "Excel"], key="sb_exp_fmt")
    export_tpl = st.sidebar.selectbox(
        t("export_template"),
        [t("export_all_tables"), t("export_summary_only")],
        key="sb_exp_tpl",
    )

    if st.sidebar.button(t("export_btn"), key="btn_export"):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_tpl == t("export_all_tables"):
            tables = [
                ("vessel_summary", df_vessel_raw),
                ("qc_productivity", df_qc),
                ("qc_operator_productivity", df_qc_operator),
                ("delay_details", df_delay),
                ("container_details", df_container),
            ]
        else:
            tables = [("vessel_summary", df_vessel_raw)]

        if export_fmt == "CSV":
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for name, df in tables:
                    if not df.empty:
                        zf.writestr(f"{name}.csv", df.to_csv(index=False))
            buf.seek(0)
            st.sidebar.download_button(
                "📥 Download ZIP", buf,
                file_name=f"tdr_export_{ts}.zip", mime="application/zip",
            )
        else:
            # S4-2: styled Excel export
            buf = _create_styled_excel(tables)
            st.sidebar.download_button(
                "📥 Download Excel", buf,
                file_name=f"tdr_export_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
