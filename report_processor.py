import logging
import threading
from pathlib import Path
import pandas as pd
import openpyxl  # For InvalidFileException and load_workbook
from datetime import date, datetime  # For type hints (removed 'time' to avoid naming conflict)
import time as time_module  # Use 'time_module' to avoid naming conflict with datetime.time
import re
from typing import Dict, List, Optional, Set, Any, Tuple

try:
    import config
    from data_extractors import DataExtractor
    from utils.excel_handler import append_df_to_excel
    from utils.file_utils import backup_file
except ImportError as e:
    logging.error(f"Import error in report_processor.py: {e}")
    raise
def _normalize_qc_name(name: str) -> str:
    """Chuẩn hóa tên QC về định dạng 'XX00'. Ví dụ: 'GC1' -> 'GC01'."""
    if not isinstance(name, str):
        return ""
    name = name.strip().upper()
    # Tách phần chữ và phần số
    letters = ''.join(re.findall(r'[A-Z]', name))
    numbers = ''.join(re.findall(r'\d', name))
    if letters and numbers:
        # Format lại số có 2 chữ số, ví dụ: 1 -> 01
        return f"{letters}{int(numbers):02d}"
    return name # Trả về tên gốc nếu không phân tích được

class ReportProcessor:
    def __init__(self, input_dir: Optional[Path] = None, output_dir: Optional[Path] = None) -> None:
        # Convert strings to Path if needed
        if isinstance(input_dir, str):
            input_dir = Path(input_dir)
        if isinstance(output_dir, str):
            output_dir = Path(output_dir)

        self.input_dir: Path = input_dir if input_dir else Path("data_input")
        self.output_dir: Path = output_dir if output_dir else Path("outputs")
        self.data_excel_dir: Path = self.output_dir / "data_excel"
        self.data_csv_dir: Path = self.output_dir / "data_csv"

        self.skipped_files_log: List[str] = []
        self.all_vessel_dfs: List[pd.DataFrame] = []
        self.all_qc_dfs: List[pd.DataFrame] = []
        self.all_qc_operator_dfs: List[pd.DataFrame] = []
        self.all_delay_dfs: List[pd.DataFrame] = []
        self.all_container_long_dfs: List[pd.DataFrame] = []
        self.processed_files_count: int = 0
        self.skipped_files_count: int = 0

        # Thread safety: prevent concurrent processing runs from corrupting shared state
        self._processing_lock: threading.Lock = threading.Lock()

        self.data_excel_dir.mkdir(parents=True, exist_ok=True)
        self.data_csv_dir.mkdir(parents=True, exist_ok=True)

    def _get_existing_processed_filenames(self) -> Set[str]:
        vessel_master_filepath: Path = self.data_excel_dir / config.VESSEL_MASTER_FILE
        existing_filenames: Set[str] = set()
        if vessel_master_filepath.exists():
            try:
                wb = openpyxl.load_workbook(vessel_master_filepath, read_only=True, data_only=True)
                if config.VESSEL_SUMMARY_SHEET in wb.sheetnames:
                    ws = wb[config.VESSEL_SUMMARY_SHEET]
                    header = [cell.value for cell in ws[1]]
                    if 'Filename' in header:
                        filename_col_idx = header.index('Filename') + 1
                        for row in ws.iter_rows(min_row=2, min_col=filename_col_idx, max_col=filename_col_idx, values_only=True):
                            if row[0]:
                                existing_filenames.add(row[0])
                wb.close()
            except Exception as e:
                logging.warning(f"Không thể đọc {config.VESSEL_MASTER_FILE} để kiểm tra file đã xử lý: {e}")
        return existing_filenames

    # THAY THẾ HOÀN TOÀN HÀM _process_single_file
    def _process_single_file(self, filepath: Path) -> bool:
        logging.info(f"=== Processing file: {filepath.name} (called from ReportProcessor) ===")
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            extractor = DataExtractor(ws, filepath)
            
            # --- BƯỚC 1: TRÍCH XUẤT TẤT CẢ DỮ LIỆU THÔ ---
            vessel_info: Dict[str, Any] = extractor.extract_vessel_info()
            initial_qc_list: List[Dict[str, Any]] = extractor.extract_qc_productivity()
            delay_list_detailed: List[Dict[str, Any]] = extractor.extract_delay_details(extractor.reference_date_for_events)
            container_list_long: List[Dict[str, Any]] = extractor.extract_container_details()

            # --- BƯỚC 2: TÍNH TOÁN CHO VESSEL SUMMARY (GIỮ NGUYÊN) ---
            vessel_info["Break Time (hrs)"] = round(vessel_info.get("Break Dis (hrs)", 0.0) + vessel_info.get("Break Load (hrs)", 0.0), 2)
            vessel_info["Net Working (hrs)"] = round(max(0, vessel_info.get("Gross Working (hrs)", 0.0) - vessel_info.get("Break Time (hrs)", 0.0)), 2)
            
            grand_total_containers: int = vessel_info.get("Grand Total Conts", 0)
            net_working_hours_vessel: float = vessel_info.get("Net Working (hrs)", 0)
            gross_working_hours_vessel: float = vessel_info.get("Gross Working (hrs)", 0)
            portstay_hours_vessel: float = vessel_info.get("Portstay (hrs)", 0)

            vessel_info["Vessel Moves/Net Hour"] = round(grand_total_containers / net_working_hours_vessel, 1) if net_working_hours_vessel > 0 and grand_total_containers > 0 else 0.0
            vessel_info["Vessel Moves/Gross Hour"] = round(grand_total_containers / gross_working_hours_vessel, 1) if gross_working_hours_vessel > 0 and grand_total_containers > 0 else 0.0
            vessel_info["Vessel Moves/Portstay Hour"] = round(grand_total_containers / portstay_hours_vessel, 1) if portstay_hours_vessel > 0 and grand_total_containers > 0 else 0.0
            
            self.all_vessel_dfs.append(pd.DataFrame([vessel_info]))

            # --- BƯỚC 3: TẠO CÁC DATAFRAME CHI TIẾT ---
            df_qc = pd.DataFrame(initial_qc_list)
            
            # Tạo qc_productivity_df (dùng delay từ bảng productivity) - Giữ nguyên
            if not df_qc.empty:
                qc_productivity_df = df_qc.copy()
                qc_productivity_df["Net working (hrs)"] = qc_productivity_df.apply(
                    lambda row: round(max(0, row.get("Gross working (hrs)", 0.0) - row.get("Delay times (hrs)", 0.0)), 2), axis=1
                )
                # Use "Net moves/h" to match QC_COLS_OUTPUT_ORDER in config
                qc_productivity_df["Net moves/h"] = qc_productivity_df.apply(
                    lambda row: round(row["Total Conts"] / row["Net working (hrs)"], 1) if row["Net working (hrs)"] > 0 else 0.0, axis=1
                )
                self.all_qc_dfs.append(qc_productivity_df)

            # <<< LOGIC MỚI VÀ ĐÃ SỬA LỖI: Tạo qc_operator_df >>>
            if not df_qc.empty:
                qc_operator_df = df_qc.copy()
                qc_operator_df['ATB'] = vessel_info.get('ATB')
                qc_operator_df['ATD'] = vessel_info.get('ATD')

                # 1. Chuyển danh sách delay chi tiết thành DataFrame
                delay_details_df = pd.DataFrame(delay_list_detailed)
                
                # 2. Tính tổng delay cho mỗi QC từ bảng chi tiết
                qc_actual_delays = pd.Series(dtype=float)
                if not delay_details_df.empty:
                    # <<< SỬA LỖI: Chuẩn hóa tên QC trước khi groupby >>>
                    delay_details_df['QC_No_Normalized'] = delay_details_df['QC No.'].apply(_normalize_qc_name)
                    qc_actual_delays = delay_details_df.groupby('QC_No_Normalized')['Duration (hrs)'].sum()

                # 3. Áp dụng tổng delay này vào bảng QC Operator
                def calculate_operator_qc_metrics(row):
                    # <<< SỬA LỖI: Chuẩn hóa tên QC trước khi tra cứu >>>
                    qc_no_normalized = _normalize_qc_name(row['QC No.'])
                    actual_delay = qc_actual_delays.get(qc_no_normalized, 0.0)
                    
                    # Đổi tên cột thành "Total Stop Time (hrs)"
                    row['Total Stop Time (hrs)'] = actual_delay 
                    row['Net working (hrs)'] = round(max(0, row.get("Gross working (hrs)", 0.0) - actual_delay), 2)
                    
                    if row["Net working (hrs)"] > 0 and row.get("Total Conts", 0) > 0:
                        row["Net moves/h"] = round(row["Total Conts"] / row["Net working (hrs)"], 1)
                    else:
                        row["Net moves/h"] = 0.0
                    return row
                
                df_qc_operator_final = qc_operator_df.apply(calculate_operator_qc_metrics, axis=1)
                self.all_qc_operator_dfs.append(df_qc_operator_final)

            self.all_delay_dfs.append(pd.DataFrame(delay_list_detailed))
            self.all_container_long_dfs.append(pd.DataFrame(container_list_long))

            logging.info(f"=== Finished processing: {filepath.name} (called from ReportProcessor) ===")
            return True

        except Exception as e:
            logging.error(f"PROCESSOR_ERROR: Error processing {filepath.name}: {e}", exc_info=True)
            return False

    def _save_skipped_log(self):
        if not self.skipped_files_log:
            return
        
        log_path = self.output_dir / "skipped_files_log.xlsx"
        df_skipped = pd.DataFrame(self.skipped_files_log)
        
        try:
            df_skipped.to_excel(log_path, index=False, sheet_name="Skipped Files Log")
            logging.info(f"Đã ghi log các file bị bỏ qua vào: {log_path}")
        except Exception as e:
            logging.error(f"Không thể ghi file log các file bị bỏ qua: {e}")

    def _process_container_data(self, df_container_long_raw: pd.DataFrame):
        if df_container_long_raw.empty:
            return pd.DataFrame(), pd.DataFrame()
        
        # Lọc dữ liệu
        temp_df_c = df_container_long_raw.copy()
        condition_total_ops = temp_df_c['OperationType'].isin([config.OP_TOTAL_DIS, config.OP_TOTAL_LOAD])
        temp_df_c = temp_df_c[~condition_total_ops]
        condition_dis_load_all_ports = (
            temp_df_c['OperationType'].isin([config.OP_DISCHARGE, config.OP_LOADING]) &
            (temp_df_c['Port'] == config.PORT_ALL))
        temp_df_c = temp_df_c[~condition_dis_load_all_ports]
        condition_shifting_all_ports = (
            temp_df_c['OperationType'].isin([config.OP_SHIFTING_DIS, config.OP_SHIFTING_LOAD]) &
            (temp_df_c['Port'] == config.PORT_ALL))
        temp_df_c = temp_df_c[~condition_shifting_all_ports]
        df_container_long_filtered = temp_df_c.reset_index(drop=True)

        # Pivot sang wide format
        df_container_wide = pd.DataFrame()
        if not df_container_long_filtered.empty:
            try:
                df_container_long_filtered['ContainerTypeSize'] = df_container_long_filtered['ContainerCategory'] + '_' + df_container_long_filtered['ContainerSize'].astype(str)
                df_container_pivot = pd.pivot_table(df_container_long_filtered,
                                                    values='Quantity',
                                                    index=config.CONTAINER_WIDE_BASE_COLS,
                                                    columns='ContainerTypeSize',
                                                    aggfunc='sum',
                                                    fill_value=0)
                df_container_wide = df_container_pivot.reset_index()
                quantity_cols_in_wide = [col for col in df_container_wide.columns if col not in config.CONTAINER_WIDE_BASE_COLS]
                if quantity_cols_in_wide:
                    df_container_wide[config.CONTAINER_WIDE_TOTAL_COLS_SUFFIX[0]] = df_container_wide[quantity_cols_in_wide].sum(axis=1)
                    df_container_wide[config.CONTAINER_WIDE_TOTAL_COLS_SUFFIX[1]] = 0
                    for col_name in quantity_cols_in_wide:
                        if "_20" in col_name:
                            df_container_wide[config.CONTAINER_WIDE_TOTAL_COLS_SUFFIX[1]] += df_container_wide[col_name] * 1
                        elif "_40" in col_name:
                            df_container_wide[config.CONTAINER_WIDE_TOTAL_COLS_SUFFIX[1]] += df_container_wide[col_name] * 2
                        elif "_45" in col_name:
                            df_container_wide[config.CONTAINER_WIDE_TOTAL_COLS_SUFFIX[1]] += df_container_wide[col_name] * 2
                
                ordered_value_cols = []
                for cat_cfg in config.CONTAINER_CATEGORIES_IN_ORDER:
                    for size_cfg in config.CONTAINER_SIZES_IN_ORDER:
                        col_name = f"{cat_cfg}_{size_cfg}"
                        if col_name in df_container_wide.columns:
                            ordered_value_cols.append(col_name)
                
                remaining_value_cols_in_wide = [col for col in quantity_cols_in_wide if col not in ordered_value_cols]
                final_container_wide_col_order = config.CONTAINER_WIDE_BASE_COLS + ordered_value_cols + remaining_value_cols_in_wide + config.CONTAINER_WIDE_TOTAL_COLS_SUFFIX
                final_container_wide_col_order = [col for col in final_container_wide_col_order if col in df_container_wide.columns]
                df_container_wide = df_container_wide[final_container_wide_col_order]
                
                df_container_wide['OpSortOrder'] = df_container_wide['OperationType'].map(config.OPERATION_TYPE_SORT_ORDER).fillna(999)
                df_container_wide = df_container_wide.sort_values(by=['Filename', 'Vessel Name', 'Voyage', 'OpSortOrder', 'Port'], ascending=[True, True, True, True, True])
                df_container_wide = df_container_wide.drop(columns=['OpSortOrder'])
            except Exception as e_pivot:
                logging.error(f"Lỗi pivot/sắp xếp container details: {e_pivot}", exc_info=True)
        
        return df_container_long_filtered, df_container_wide

    def _save_csv_files(self, final_dataframes: dict):
        logging.info("Bắt đầu chuyển đổi tất cả master data sang CSV...")
        csv_map = {
            "vessel_summary.csv": config.VESSEL_MASTER_FILE,
            "qc_productivity.csv": config.QC_MASTER_FILE,
            "delay_details.csv": config.DELAY_MASTER_FILE,
            "container_details_wide.csv": config.CONTAINER_MASTER_WIDE_FILE,
            "container_details_long.csv": config.CONTAINER_MASTER_LONG_FILE,
            # <<< THÊM FILE MỚI VÀO MAP
            "qc_operator_productivity.csv": config.QC_OPERATOR_MASTER_FILE
        }
        
        for csv_name, excel_name in csv_map.items():
            if excel_name in final_dataframes and final_dataframes[excel_name] is not None:
                df_to_write_csv = final_dataframes[excel_name]
                csv_path = self.data_csv_dir / csv_name
                df_to_write_csv.to_csv(csv_path, index=False, encoding='utf-8-sig')
                logging.info(f"Đã xuất file CSV: {csv_name} với {len(df_to_write_csv)} dòng.")
            else:
                logging.info(f"Bỏ qua xuất file CSV '{csv_name}' vì không có dữ liệu tương ứng.")

    def _aggregate_and_save_data(self) -> str:
        if not self.all_vessel_dfs and not self.all_qc_dfs and not self.all_delay_dfs and not self.all_container_long_dfs:
            if self.skipped_files_count > 0 and self.processed_files_count == 0:
                return "Tất cả file đã chọn đều đã được xử lý trước đó."
            return "Không có dữ liệu mới hợp lệ nào được trích xuất."

        # --- Bước 1: Concat tất cả dữ liệu mới ---
        df_vessel_new = pd.concat(self.all_vessel_dfs, ignore_index=True) if self.all_vessel_dfs else pd.DataFrame()
        df_qc_new = pd.concat(self.all_qc_dfs, ignore_index=True) if self.all_qc_dfs else pd.DataFrame()
        df_qc_operator_new = pd.concat(self.all_qc_operator_dfs, ignore_index=True) if self.all_qc_operator_dfs else pd.DataFrame()
        df_delay_new = pd.concat(self.all_delay_dfs, ignore_index=True) if self.all_delay_dfs else pd.DataFrame()
        df_container_long_new = pd.concat(self.all_container_long_dfs, ignore_index=True) if self.all_container_long_dfs else pd.DataFrame()

        # --- Bước 2: Đổi tên và lọc cột cho từng DataFrame ---
        # Safely select columns - only include columns that actually exist in the DataFrame
        def safe_select_columns(df, col_renames, col_order):
            """Safely rename and select columns from DataFrame"""
            if df.empty:
                return df
            
            # Apply column renames (if any exist)
            if col_renames:
                df = df.rename(columns=col_renames)
            
            # Filter to only include columns that exist in the DataFrame
            existing_cols = [col for col in col_order if col in df.columns]
            if not existing_cols:
                logging.warning(f"No columns from order list found in DataFrame. Columns available: {df.columns.tolist()}")
                return df
            
            return df[existing_cols]
        
        final_df_vessel = safe_select_columns(df_vessel_new, config.VESSEL_COL_RENAMES_OUTPUT, config.VESSEL_COLS_OUTPUT_ORDER)
        final_df_qc = safe_select_columns(df_qc_new, config.QC_COL_RENAMES_OUTPUT, config.QC_COLS_OUTPUT_ORDER)
        
        # <<< THAY ĐỔI TÊN CỘT VÀ THỨ TỰ CỘT CHO QC OPERATOR >>>
        # Bỏ cột "Delay times (hrs)" cũ và thêm cột "Total Stop Time (hrs)" mới
        QC_OPERATOR_COLS_OUTPUT_ORDER_NEW = [
            'Filename', 'Vessel Name', 'Voyage', 'ATB', 'ATD', 'QC No.', 'Start Time', 'End Time',
            'Gross working (hrs)', 'Total Stop Time (hrs)', 'Net working (hrs)', 'Discharge Conts',
            'Load Conts', 'Shifting Conts', 'Total Conts', 'Gross moves/h', 'Net moves/h'
        ]
        # Đổi tên cột "Delay times (hrs)" thành "Total Stop Time (hrs)"
        df_qc_operator_new = df_qc_operator_new.drop(columns=['Delay times (hrs)'], errors='ignore')
        final_df_qc_operator = safe_select_columns(df_qc_operator_new, config.QC_OPERATOR_COL_RENAMES_OUTPUT, QC_OPERATOR_COLS_OUTPUT_ORDER_NEW)
        # <<< KẾT THÚC THAY ĐỔI >>>

        final_df_delay = safe_select_columns(df_delay_new, {}, config.DELAY_DETAILS_COLS_OUTPUT_ORDER)
        final_df_container_long, final_df_container_wide = self._process_container_data(df_container_long_new)

        # --- Bước 3: Ghi file Excel ---
        output_map_excel = {
            config.VESSEL_MASTER_FILE: (final_df_vessel, config.VESSEL_SUMMARY_SHEET, {"datetime_cols_to_format": config.VESSEL_DATETIME_COLS_FORMAT, "index": False}),
            config.QC_MASTER_FILE: (final_df_qc, config.QC_PRODUCTIVITY_SHEET, {"time_cols_to_format": config.QC_TIME_COLS_FORMAT, "index": False}),
            config.QC_OPERATOR_MASTER_FILE: (final_df_qc_operator, config.QC_OPERATOR_PRODUCTIVITY_SHEET, {"index": False}),
            config.DELAY_MASTER_FILE: (final_df_delay, config.DELAY_DETAILS_SHEET, {"datetime_cols_to_format": config.DELAY_DATETIME_COLS_FORMAT, "index": False}),
            config.CONTAINER_MASTER_WIDE_FILE: (final_df_container_wide, config.CONTAINER_SUMMARY_WIDE_SHEET, {"index": False}),
            config.CONTAINER_MASTER_LONG_FILE: (final_df_container_long, config.CONTAINER_SUMMARY_LONG_SHEET, {"index": False})
        }
        
        final_dataframes_for_csv = {}
        for filename, (df, sheet_name, kwargs) in output_map_excel.items():
            if df is not None and not df.empty:
                file_path = self.data_excel_dir / filename
                backup_file(file_path)
                final_df = append_df_to_excel(file_path, df, sheet_name, **kwargs)
                if final_df is not None:
                    final_dataframes_for_csv[filename] = final_df

        # --- Bước 4: Ghi file CSV ---
        self._save_csv_files(final_dataframes_for_csv)
        self._save_skipped_log()

        msg = f"Xử lý hoàn tất!\nĐã cập nhật dữ liệu vào thư mục:\n{self.data_excel_dir}"
        if self.skipped_files_count > 0:
            msg += f"\nĐã bỏ qua {self.skipped_files_count} file (xem chi tiết trong skipped_files_log.xlsx)."
        return msg

    def process_tdr_files(self, input_file_paths: list[Path], update_status_callback=None,
                          update_progress_callback=None, overwrite: bool = False) -> dict:
        # Acquire lock to prevent concurrent processing runs from corrupting shared state
        if not self._processing_lock.acquire(blocking=False):
            logging.warning("ReportProcessor: Processing already in progress. Skipping concurrent request.")
            return {
                "message": "Processing already in progress. Please wait.",
                "time_taken": 0.0,
                "processed_count": 0,
                "skipped_count": len(input_file_paths)
            }
        try:
            return self._process_tdr_files_internal(
                input_file_paths, update_status_callback, update_progress_callback, overwrite
            )
        finally:
            self._processing_lock.release()

    def _process_tdr_files_internal(self, input_file_paths: list[Path], update_status_callback=None,
                                     update_progress_callback=None, overwrite: bool = False) -> dict:
        start_total_time = time_module.perf_counter()
        self.all_vessel_dfs.clear()
        self.all_qc_dfs.clear()
        self.all_qc_operator_dfs.clear()  # Bug fix: was missing in refactored method
        self.all_delay_dfs.clear()
        self.all_container_long_dfs.clear()
        self.processed_files_count = 0
        self.skipped_files_count = 0
        self.skipped_files_log.clear()

        existing_filenames = set()
        if not overwrite:
            existing_filenames = self._get_existing_processed_filenames()
            logging.info(f"Chế độ 'Chỉ xử lý file mới'. Tìm thấy {len(existing_filenames)} file đã có trong master data.")
        else:
            logging.info("Chế độ 'Xử lý lại' được bật. Sẽ xử lý tất cả các file được chọn.")

        total_files = len(input_file_paths)

        if update_progress_callback:
            update_progress_callback(0, total_files)

        for idx, filepath_obj in enumerate(input_file_paths):
            if update_status_callback:
                update_status_callback(f"Kiểm tra file: {filepath_obj.name}")

            if not overwrite and filepath_obj.name in existing_filenames:
                logging.info(f"Bỏ qua file đã xử lý: {filepath_obj.name}")
                self.skipped_files_count += 1
                self.skipped_files_log.append({
                    "filename": filepath_obj.name,
                    "reason": "Đã được xử lý trước đó",
                    "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                if update_progress_callback:
                    update_progress_callback(idx + 1, total_files)
                continue

            if update_status_callback:
                update_status_callback(f"✅Processing...: {filepath_obj.name} ({idx+1}/{total_files})")

            success = self._process_single_file(filepath_obj)
            if success:
                self.processed_files_count += 1
            else:
                self.skipped_files_count += 1
                self.skipped_files_log.append({
                    "filename": filepath_obj.name,
                    "reason": "Lỗi trong quá trình xử lý (xem tdr_processor.log)",
                    "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })

            if update_progress_callback:
                update_progress_callback(idx + 1, total_files)
        
        try:
            result_message = self._aggregate_and_save_data()
            end_total_time = time_module.perf_counter()
            time_taken = end_total_time - start_total_time

            logging.info(f"Tổng thời gian xử lý {self.processed_files_count} file mới: {time_taken:.2f} giây")

            return {
                "message": result_message,
                "time_taken": time_taken,
                "processed_count": self.processed_files_count,
                "skipped_count": self.skipped_files_count
            }

        except Exception as e_agg:
            end_total_time = time_module.perf_counter()
            time_taken = end_total_time - start_total_time
            logging.error(f"Lỗi nghiêm trọng khi tổng hợp hoặc lưu dữ liệu: {e_agg}", exc_info=True)
            logging.info(f"Tổng thời gian xử lý (kể cả lỗi): {time_taken:.2f} giây")
            return {
                "message": f"Lỗi nghiêm trọng khi tổng hợp/lưu dữ liệu: {e_agg}",
                "time_taken": time_taken,
                "processed_count": self.processed_files_count,
                "skipped_count": self.skipped_files_count
            }