"""
utils/datetime_utils.py - Common datetime parsing and extraction utilities for TDR Processor

This module provides reusable functions for extracting and parsing datetime values from
Excel files, reducing code duplication and improving consistency across data extraction.

Type Hints: Fully type-hinted for better IDE support and type checking with mypy.
"""

import logging
from typing import Optional, Any
from datetime import datetime, date, time, timedelta
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

try:
    import config
    from utils.excel_utils import parse_excel_datetime, excel_to_time, col_letter_to_index
except ImportError as e:
    logging.warning(f"Import error in datetime_utils.py: {e}")
    # Fallback imports if running in different context
    pass


def extract_datetime_from_cell(
    worksheet: Worksheet,
    row: int,
    column: int,
    reference_date: Optional[date] = None,
    context: str = ""
) -> Optional[datetime]:
    """
    Extract and parse a datetime value from an Excel cell.
    
    Reads the raw value from an Excel cell and parses it into a datetime object using
    the parse_excel_datetime function. If a reference date is provided, it's used for
    resolving ambiguous datetime values.
    
    Args:
        worksheet: Excel worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        reference_date: Optional reference date for parsing (used if cell contains only time)
        context: Context string for logging purposes (e.g., "ATD", "Report Date")
    
    Returns:
        Parsed datetime object or None if unable to parse
    
    Example:
        >>> from openpyxl import load_workbook
        >>> from datetime import date
        >>> wb = load_workbook('file.xlsx')
        >>> ws = wb.active
        >>> dt = extract_datetime_from_cell(ws, 10, 5, date(2023, 7, 15), "ATD")
        >>> print(dt)
        2023-07-15 14:30:00
    """
    try:
        raw_value = worksheet.cell(row=row, column=column).value
        
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            logging.debug(f"extract_datetime_from_cell ({context}): Empty cell at row {row}, col {column}")
            return None
        
        parsed_datetime = parse_excel_datetime(raw_value, reference_date, False, context)
        return parsed_datetime
    
    except Exception as e:
        logging.warning(f"extract_datetime_from_cell ({context}): Error parsing cell ({row}, {column}): {e}")
        return None


def extract_date_from_cell(
    worksheet: Worksheet,
    row: int,
    column: int,
    context: str = ""
) -> Optional[date]:
    """
    Extract and parse a date value from an Excel cell.
    
    Reads the raw value from an Excel cell and parses it into a date object. This function
    extracts only the date portion, ignoring any time component.
    
    Args:
        worksheet: Excel worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        context: Context string for logging purposes (e.g., "Report Date")
    
    Returns:
        Parsed date object or None if unable to parse
    
    Example:
        >>> date_val = extract_date_from_cell(ws, 5, 3, "Report Date")
        >>> print(date_val)
        2023-07-15
    """
    try:
        datetime_result = extract_datetime_from_cell(worksheet, row, column, None, context)
        if isinstance(datetime_result, datetime):
            return datetime_result.date()
        return None
    except Exception as e:
        logging.warning(f"extract_date_from_cell ({context}): Error at row {row}, col {column}: {e}")
        return None


def extract_time_from_cell(
    worksheet: Worksheet,
    row: int,
    column: int,
    context: str = ""
) -> Optional[time]:
    """
    Extract and parse a time value from an Excel cell.
    
    Reads the raw value from an Excel cell and parses it into a time object.
    
    Args:
        worksheet: Excel worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        context: Context string for logging purposes (e.g., "Start Time")
    
    Returns:
        Parsed time object or None if unable to parse
    
    Example:
        >>> time_val = extract_time_from_cell(ws, 10, 4, "Start Time")
        >>> print(time_val)
        14:30:00
    """
    try:
        raw_value = worksheet.cell(row=row, column=column).value
        
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            return None
        
        parsed_time = excel_to_time(raw_value)
        return parsed_time
    except Exception as e:
        logging.warning(f"extract_time_from_cell ({context}): Error parsing cell ({row}, {column}): {e}")
        return None


def parse_datetime_range(
    start_value: Any,
    end_value: Any,
    reference_date: Optional[date] = None,
    context: str = ""
) -> tuple[Optional[datetime], Optional[datetime]]:
    """
    Parse a start and end datetime, handling multi-day scenarios.
    
    Converts raw Excel values into datetime objects. If the end time is earlier than
    the start time on the same date, assumes the end time is on the next day.
    
    Args:
        start_value: Raw start datetime value from Excel
        end_value: Raw end datetime value from Excel
        reference_date: Reference date to use for parsing
        context: Context string for logging purposes
    
    Returns:
        Tuple of (start_datetime, end_datetime) or (None, None) if parsing fails
    
    Example:
        >>> start = extract_time_from_cell(ws, 10, 2)  # 14:30
        >>> end = extract_time_from_cell(ws, 10, 3)    # 02:30 (next day)
        >>> start_dt, end_dt = parse_datetime_range(start, end, date(2023, 7, 15))
        >>> print(start_dt)
        2023-07-15 14:30:00
        >>> print(end_dt)
        2023-07-16 02:30:00
    """
    try:
        start_datetime = parse_excel_datetime(start_value, reference_date, False, f"{context} - Start")
        end_datetime = parse_excel_datetime(end_value, reference_date, False, f"{context} - End")
        
        # If both are valid datetimes and end is before start, add a day to end
        if (isinstance(start_datetime, datetime) and isinstance(end_datetime, datetime) 
            and end_datetime < start_datetime):
            end_datetime += timedelta(days=1)
            logging.debug(f"parse_datetime_range ({context}): End time was earlier than start, added 1 day")
        
        return start_datetime, end_datetime
    
    except Exception as e:
        logging.warning(f"parse_datetime_range ({context}): Error parsing datetime range: {e}")
        return None, None


def calculate_duration_hours(
    start_datetime: Optional[datetime],
    end_datetime: Optional[datetime],
    context: str = ""
) -> float:
    """
    Calculate duration in hours between two datetime objects.
    
    Computes the time difference between start and end datetimes and converts it to hours.
    
    Args:
        start_datetime: Start datetime
        end_datetime: End datetime
        context: Context string for logging purposes
    
    Returns:
        Duration in hours, rounded to 2 decimal places. Returns 0.0 if inputs are invalid.
    
    Example:
        >>> from datetime import datetime
        >>> start = datetime(2023, 7, 15, 14, 30)
        >>> end = datetime(2023, 7, 15, 22, 45)
        >>> hours = calculate_duration_hours(start, end, "Working Period")
        >>> print(hours)
        8.25
    """
    try:
        if not (isinstance(start_datetime, datetime) and isinstance(end_datetime, datetime)):
            return 0.0
        
        if end_datetime < start_datetime:
            logging.warning(f"calculate_duration_hours ({context}): End time before start time")
            return 0.0
        
        time_delta = end_datetime - start_datetime
        hours = round(time_delta.total_seconds() / 3600.0, 2)
        return max(0.0, hours)  # Ensure non-negative
    
    except Exception as e:
        logging.warning(f"calculate_duration_hours ({context}): Error calculating duration: {e}")
        return 0.0


def shift_datetime_to_reference_date(
    time_value: Optional[time],
    reference_date: Optional[date],
    context: str = ""
) -> Optional[datetime]:
    """
    Combine a time value with a reference date to create a datetime.
    
    Useful for converting standalone time values (from Excel cells) into full datetime
    objects using a reference date (e.g., the vessel's report date or first operation date).
    
    Args:
        time_value: Time object to shift
        reference_date: Date to combine with the time
        context: Context string for logging purposes
    
    Returns:
        Combined datetime object or None if inputs are invalid
    
    Example:
        >>> from datetime import time, date
        >>> t = time(14, 30, 0)
        >>> d = date(2023, 7, 15)
        >>> dt = shift_datetime_to_reference_date(t, d, "ATD")
        >>> print(dt)
        2023-07-15 14:30:00
    """
    try:
        if not isinstance(time_value, time) or not isinstance(reference_date, date):
            logging.debug(f"shift_datetime_to_reference_date ({context}): Invalid input types")
            return None
        
        combined_datetime = datetime.combine(reference_date, time_value)
        logging.debug(f"shift_datetime_to_reference_date ({context}): Combined {time_value} with {reference_date}")
        return combined_datetime
    
    except Exception as e:
        logging.warning(f"shift_datetime_to_reference_date ({context}): Error combining datetime: {e}")
        return None
