"""
utils/excel_utils.py - Excel file parsing and manipulation utilities for TDR Processor

This module provides utility functions for working with Excel files, including:
- Column letter to index conversion
- Timedelta to hours conversion
- Error code classification
- Label and cell value finding
- DateTime parsing from Excel formats
- Time duration parsing

Type Hints: Fully type-hinted for mypy compatibility.
"""

import logging
import re
from datetime import datetime, time, date, timedelta
import pandas as pd
from openpyxl.utils.datetime import from_excel as openpyxl_from_excel
from openpyxl.utils import get_column_letter
try:
    import config # Giả sử config.py ở thư mục gốc của project
except ImportError:
    print("LỖI: Không tìm thấy config.py trong excel_utils.")
    # Mock config nếu cần cho testing riêng lẻ module này
    class ConfigMockExcel:
        MAX_SEARCH_ROWS_DEFAULT = 70
        PARTIAL_MATCH_DEFAULT = True
        DELAY_ERROR_UNKNOWN_TYPE = "Unknown"
        DELAY_ERROR_CODE_CLASSIFICATION = {
            "Terminal Convenience": list('abcdefgh'),
            "Non-Terminal Convenience": list('ijklmn')
        }
        DELAY_ERROR_FORCE_MAJEURE_KEYWORDS = ["weather", "bất khả kháng", "thời tiết"]
    config = ConfigMockExcel()


# Các hàm helper đã có trong extract_updated.py sẽ được chuyển vào đây
# và có thể được nhóm vào một class ExcelUtils sau này.

def col_letter_to_index(letter: str) -> int | None:
    """
    Convert Excel column letter to column index number.
    
    Converts column letters (A, B, Z, AA, AB, etc.) to their corresponding
    column numbers (1, 2, 26, 27, 28, etc.). Supports multi-letter columns.
    
    Args:
        letter: Excel column letter (e.g., "A", "AA", "Z")
    
    Returns:
        Column index (1-indexed) or None if letter is invalid
    
    Example:
        >>> col_letter_to_index("A")
        1
        >>> col_letter_to_index("Z")
        26
        >>> col_letter_to_index("AA")
        27
    """
    if not isinstance(letter, str) or not letter:
        return None
    num = 0
    letter = letter.upper()
    for char_val in letter:
        if 'A' <= char_val <= 'Z': # Kiểm tra ký tự hợp lệ
            num = num * 26 + (ord(char_val) - ord('A')) + 1
        else:
            logging.debug(f"col_letter_to_index: Invalid character '{char_val}' in input '{letter}'.")
            return None # Nếu gặp ký tự không phải chữ cái, trả về None
    return num

def timedelta_to_hours(td: timedelta) -> float:
    """
    Convert timedelta object to hours with 2 decimal precision.
    
    Converts a timedelta object into a float representing hours, rounded
    to 2 decimal places. Returns 0.0 for non-timedelta inputs.
    
    Args:
        td: timedelta object to convert
    
    Returns:
        Hours as float, rounded to 2 decimal places
    
    Example:
        >>> from datetime import timedelta
        >>> timedelta_to_hours(timedelta(hours=2, minutes=30))
        2.5
        >>> timedelta_to_hours(timedelta(hours=1, minutes=15))
        1.25
    """
    if isinstance(td, timedelta):
        return round(td.total_seconds() / 3600.0, 2)
    return 0.0

def classify_error_code(error_code_and_remark: str | None) -> tuple[str | None, str]:
    """
    Classify delay error code based on predefined categories.
    
    Extracts error code (first letter) from error code/remark text and classifies
    it as "Terminal Convenience", "Non-Terminal Convenience", "Other/Force Majeure",
    or "Unknown".
    
    Args:
        error_code_and_remark: String containing error code and/or remark text
    
    Returns:
        Tuple of (error_code, classification) where:
        - error_code: Extracted single-letter code or None
        - classification: One of the predefined categories
    
    Example:
        >>> classify_error_code("A - Terminal issue")
        ('a', 'Terminal Convenience')
        >>> classify_error_code("weather delay")
        (None, 'Other/Force Majeure')
        >>> classify_error_code(None)
        (None, 'Unknown')
    """
    if not error_code_and_remark or pd.isna(error_code_and_remark):
        return None, config.DELAY_ERROR_UNKNOWN_TYPE
    text = str(error_code_and_remark).strip().lower()
    error_code_match = re.match(r"^([a-zA-Z])(?:[ \-(]|$)", text)
    error_code = error_code_match.group(1).lower() if error_code_match else None

    if error_code:
        if error_code in config.DELAY_ERROR_CODE_CLASSIFICATION.get("Terminal Convenience", []):
            return error_code, "Terminal Convenience"
        elif error_code in config.DELAY_ERROR_CODE_CLASSIFICATION.get("Non-Terminal Convenience", []):
            return error_code, "Non-Terminal Convenience"

    if any(kw in text for kw in config.DELAY_ERROR_FORCE_MAJEURE_KEYWORDS):
        return error_code, "Other/Force Majeure"
    return error_code, config.DELAY_ERROR_UNKNOWN_TYPE

def find_label_row_col(ws, label_text: str, search_cols: tuple[int, int] = (1, 20),
                       start_row: int = 1, max_search_rows: int | None = None,
                       partial_match: bool | None = None, specific_col_letter: str | None = None) -> tuple[int | None, int | None]:
    """
    Find row and column indices of a label in an Excel worksheet.
    
    Searches a worksheet for a cell containing specified label text within a
    defined search range. Can optionally restrict to a specific column and
    supports both exact and partial matches.
    
    Args:
        ws: Openpyxl worksheet object
        label_text: Text to search for
        search_cols: Tuple of (start_col, end_col) indices for search range
        start_row: Row number to start searching from (1-indexed)
        max_search_rows: Maximum rows to search from start_row. None uses default.
        partial_match: If True, partial text match suffices. None uses default.
        specific_col_letter: Optional column letter to restrict search to specific column
    
    Returns:
        Tuple of (row_index, col_index) if found, (None, None) if not found.
        Indices are 1-based to match Excel convention.
    
    Example:
        >>> from openpyxl import load_workbook
        >>> ws = load_workbook("file.xlsx").active
        >>> row, col = find_label_row_col(ws, "Vessel Name")
        >>> if row:
        ...     print(f"Found at row {row}, column {col}")
    """
    label_text_lower = label_text.lower()
    col_to_search_start = search_cols[0]
    col_to_search_end = search_cols[1]

    if specific_col_letter:
        col_idx = col_letter_to_index(specific_col_letter)
        if col_idx:
            col_to_search_start, col_to_search_end = col_idx, col_idx
        else:
            logging.warning(f"find_label: Invalid col_letter '{specific_col_letter}' for '{label_text}'. Defaulting search range.")

    _max_search_rows = max_search_rows if max_search_rows is not None else config.MAX_SEARCH_ROWS_DEFAULT
    _partial_match = partial_match if partial_match is not None else config.PARTIAL_MATCH_DEFAULT

    effective_max_rows = ws.max_row + 1
    end_row_limit = start_row + _max_search_rows
    end_row_to_search = min(effective_max_rows, end_row_limit)
    if end_row_to_search <= start_row and ws.max_row >= start_row:
        end_row_to_search = ws.max_row + 1

    for r in range(start_row, end_row_to_search):
        for c in range(col_to_search_start, min(ws.max_column + 1, col_to_search_end + 1)):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                cell_val_str_lower = str(cell.value).lower().strip()
                normalized_cell_val = ' '.join(cell_val_str_lower.split())
                normalized_label_text = ' '.join(label_text_lower.split())
                if _partial_match:
                    if label_text_lower in cell_val_str_lower:
                        return r, c
                else:
                    if normalized_label_text == normalized_cell_val:
                        return r, c
    logging.info(f"find_label: Not found '{label_text}' (Cols {get_column_letter(col_to_search_start)}-{get_column_letter(col_to_search_end)}, Rows {start_row}-{end_row_to_search-1}, partial={_partial_match}).")
    return None, None

def excel_to_time(excel_time_val) -> time | None:
    """
    Convert Excel time value to Python time object.
    
    Handles multiple time representations: native time objects, datetime objects,
    Excel serial numbers (as fractions of a day), or complete Excel date+time values.
    
    Args:
        excel_time_val: Time value from Excel cell (time, datetime, int, float, or other)
    
    Returns:
        time object or None if conversion fails
    
    Example:
        >>> excel_to_time(time(14, 30, 0))
        datetime.time(14, 30)
        >>> excel_to_time(0.5)  # Excel 0.5 = 12 hours
        datetime.time(12, 0)
        >>> excel_to_time("invalid")
        None
    """
    # ... (Giữ nguyên code hàm này từ extract_updated.py)
    if isinstance(excel_time_val, time): return excel_time_val
    if isinstance(excel_time_val, datetime): return excel_time_val.time()
    if isinstance(excel_time_val, (int, float)):
        if 0 <= excel_time_val < 1: # Excel time as fraction of a day
            total_seconds = int(excel_time_val * 24 * 60 * 60)
            h, rem = divmod(total_seconds, 3600)
            m, s = divmod(rem, 60)
            return time(h, m, s)
        else: # Excel serial number for date+time, try to convert
            try:
                return openpyxl_from_excel(excel_time_val).time()
            except (ValueError, TypeError):
                logging.warning(f"excel_to_time: Cannot convert numeric value '{excel_time_val}' to time.")
                return None
    logging.debug(f"excel_to_time: Unsupported type '{type(excel_time_val)}' for value '{excel_time_val}'")
    return None


def _parse_time_string_with_strptime(time_str: str, context: str = "") -> time | None:
    """
    Parse time string using multiple format attempts.
    
    Attempts to parse a time string using common time formats including
    24-hour (HH:MM:SS, HH:MM) and 12-hour (with AM/PM) formats.
    
    Args:
        time_str: Time string to parse (e.g., "14:30:00", "2:30 PM")
        context: Optional context string for logging
    
    Returns:
        time object if parsing succeeds, None otherwise
    
    Example:
        >>> _parse_time_string_with_strptime("14:30:00")
        datetime.time(14, 30)
        >>> _parse_time_string_with_strptime("2:30 PM")
        datetime.time(14, 30)
    """
    log_prefix = f"_parse_time_s ({context})" if context else "_parse_time_s"
    # Thêm format cho AM/PM
    time_formats = [
        '%H:%M:%S', '%H:%M',
        '%I:%M:%S %p', '%I:%M %p' # Hỗ trợ 12-hour format với AM/PM
    ]
    for fmt_t in time_formats:
        try:
            return datetime.strptime(time_str.upper(), fmt_t).time() # .upper() để AM/PM thành AM/PM
        except ValueError:
            continue
    logging.warning(f"{log_prefix}: Could not parse time_str: '{time_str}' using formats {time_formats}")
    return None

def _parse_date_string_with_strptime(date_str: str, year_to_use: int, context: str = "") -> date | None:
    """
    Parse date string with fallback to year_to_use if not present.
    
    Attempts to parse a date string using multiple date formats. If parsing
    succeeds but contains no year, falls back to year_to_use parameter.
    
    Args:
        date_str: Date string to parse (e.g., "15/07/2023", "15-Jul-2023", "Jul 15, 2023")
        year_to_use: Year to use as fallback if date_str contains no year
        context: Optional context string for logging
    
    Returns:
        date object if parsing succeeds, None otherwise
    
    Example:
        >>> _parse_date_string_with_strptime("15/07/2023", 2023)
        datetime.date(2023, 7, 15)
        >>> _parse_date_string_with_strptime("15-Jul", 2023)
        datetime.date(2023, 7, 15)
    """
    log_prefix = f"_parse_date_s ({context})" if context else "_parse_date_s"
    # Thêm các format có tên tháng
    date_formats_with_year = [
        '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d',
        '%d/%m/%y', '%m/%d/%y', '%Y-%m-%d',
        '%d-%b-%Y', '%d-%B-%Y', # Ví dụ: 10-Jul-2023, 10-July-2023
        '%b %d, %Y', '%B %d, %Y' # Ví dụ: Jul 10, 2023, July 10, 2023
    ]
    for fmt_d in date_formats_with_year:
        try:
            return datetime.strptime(date_str, fmt_d).date()
        except ValueError:
            continue

    date_formats_no_year = [
        '%m/%d',  # ĐƯA LÊN TRƯỚC
        '%d/%m',
        '%d-%b', '%d-%B', 
        '%b %d', '%B %d'
    ]
    for fmt_d_no_year in date_formats_no_year:
        try:
            # Xử lý locale cho tên tháng nếu cần, nhưng strptime thường xử lý tiếng Anh mặc định
            return datetime.strptime(date_str, fmt_d_no_year).replace(year=year_to_use).date()
        except ValueError:
            continue
    logging.warning(f"{log_prefix}: Could not parse date_str: '{date_str}' using formats.")
    return None

def _parse_datetime_from_string_logic(dt_s: str, date_ref: date | None, is_date_only: bool, ctx: str = "") -> datetime | date | None:
    """
    Parse datetime string with flexible format support and reference date fallback.
    
    Handles complete datetime strings, date-only strings, or time-only strings.
    Supports multiple date formats (DD/MM/YYYY, YYYY-MM-DD, etc.), 24-hour and
    12-hour (AM/PM) time formats, and can combine time-only values with a reference date.
    
    Args:
        dt_s: Datetime string to parse (e.g., "2023-07-15 14:30:00", "15/07/2023", "14:30")
        date_ref: Reference date for time-only inputs or fallback scenarios
        is_date_only: If True, expect only date (no time) and ignore time parsing
        ctx: Optional context string for logging
    
    Returns:
        datetime, date, or None depending on input and is_date_only flag
    
    Example:
        >>> _parse_datetime_from_string_logic("2023-07-15 14:30:00", None, False)
        datetime.datetime(2023, 7, 15, 14, 30)
        >>> ref_date = date(2023, 7, 15)
        >>> _parse_datetime_from_string_logic("14:30", ref_date, False)
        datetime.datetime(2023, 7, 15, 14, 30)
    """
    log_p = f"_parse_dt_str ({ctx})" if ctx else "_parse_dt_str"
    dt_s_processed = dt_s.strip()
    dt_s_processed = re.sub(r"thg\s*\.?\s*", "-", dt_s_processed, flags=re.IGNORECASE) # "thg 7" -> " -7" (cần xử lý thêm)
                                                                                    # Tốt hơn là thay "thg " bằng tên tháng tiếng Anh nếu có thể
    dt_s_processed = dt_s_processed.replace("ngày", "").strip()
    # Thêm các chuẩn hóa khác nếu cần

    # Ưu tiên các định dạng datetime đầy đủ trước
    datetime_formats_full = [
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
        '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
        '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M',
        '%d-%b-%Y %H:%M:%S', '%d-%b-%Y %H:%M', # 10-Jul-2023 14:30
        '%d-%B-%Y %H:%M:%S', '%d-%B-%Y %H:%M',
        # Hỗ trợ AM/PM
        '%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y %I:%M %p',
        '%m/%d/%Y %I:%M:%S %p', '%m/%d/%Y %I:%M %p',
        '%Y-%m-%d %I:%M:%S %p', '%Y-%m-%d %I:%M %p',
        '%d-%b-%Y %I:%M:%S %p', '%d-%b-%Y %I:%M %p',
    ]
    if not is_date_only:
        for fmt in datetime_formats_full:
            try:
                # Chuyển AM/PM thành chữ hoa để strptime xử lý đúng
                parsed_dt = datetime.strptime(dt_s_processed.upper().replace(" SA", " AM").replace(" CH", " PM"), fmt)
                logging.info(f"{log_p} (SUCCESS - full format): Parsed '{dt_s_processed}' to {parsed_dt}")
                return parsed_dt
            except ValueError:
                pass

    # Nếu chỉ là ngày
    if is_date_only:
        year_to_use_for_date = (date_ref.year if isinstance(date_ref, (date, datetime)) else datetime.now().year)
        parsed_date_only = _parse_date_string_with_strptime(dt_s_processed, year_to_use_for_date, ctx + "_date_only_mode")
        if parsed_date_only:
            logging.info(f"{log_p} (SUCCESS - date only mode): Parsed '{dt_s_processed}' to {parsed_date_only}")
            return parsed_date_only

    # Xử lý trường hợp chỉ có giờ "HH:MM" hoặc "HH:MM:SS" (có hoặc không có AM/PM) và cần kết hợp với date_ref
    if not is_date_only and date_ref and isinstance(date_ref, date):
        # Regex này có thể không bắt được AM/PM, _parse_time_string_with_strptime sẽ xử lý
        if re.fullmatch(r"^\d{1,2}:\d{1,2}(:\d{1,2})?(\s*(AM|PM))?$", dt_s_processed, re.IGNORECASE):
            parsed_time_only = _parse_time_string_with_strptime(dt_s_processed, ctx + "_time_only_with_ref")
            if parsed_time_only:
                combined_dt = datetime.combine(date_ref, parsed_time_only)
                logging.info(f"{log_p} (SUCCESS - time-only with ref_date): Parsed '{dt_s_processed}' with ref_date '{date_ref}' to {combined_dt}")
                return combined_dt

    # Các trường hợp fallback phức tạp hơn (ví dụ: "HH:MM - DD/MM/YYYY")
    # Đoạn code này có thể cần được xem xét lại cẩn thận hơn để tránh false positive
    # hoặc có thể đơn giản hóa nếu các định dạng đầu vào không quá đa dạng.
    # Hiện tại, nó có thể không hoạt động đúng cho mọi trường hợp.
    # Ví dụ: "10:30 - 25/12"
    time_s, date_s_fallback = None, None
    m_dt_fb = re.match(r"(\d{1,2}:\d{1,2}(:\d{1,2})?(\s*(?:AM|PM))?)\s*(?:-\s*|,?\s+)?(.*)", dt_s_processed, re.IGNORECASE)
    if m_dt_fb:
        time_s = m_dt_fb.group(1).strip()
        date_s_fallback = m_dt_fb.group(4).strip() # Phần còn lại là date string

        parsed_time_fb = _parse_time_string_with_strptime(time_s, ctx + "_fb_time")
        if parsed_time_fb:
            year_use_fb = (date_ref.year if isinstance(date_ref, (date, datetime)) else datetime.now().year)
            parsed_date_fb = None
            if date_s_fallback: # Nếu có phần date string
                parsed_date_fb = _parse_date_string_with_strptime(date_s_fallback, year_use_fb, ctx + "_fb_date")
            
            if parsed_date_fb:
                return datetime.combine(parsed_date_fb, parsed_time_fb)
            elif date_ref and isinstance(date_ref, date): # Nếu không parse được date_s_fallback nhưng có date_ref
                return datetime.combine(date_ref, parsed_time_fb)

    logging.warning(f"{log_p} (fallback or unhandled): Could not parse datetime/date from: '{dt_s}'")
    return None


def parse_excel_datetime(dt_val, date_val_for_time: date | None = None, is_just_date: bool = False, context: str = "") -> datetime | date | None:
    log_p = f"parse_dt ({context})" if context else "parse_dt"
    if dt_val == 0.75: # Chỉ debug cho trường hợp này
        print(f"\nDEBUG_FUNC: parse_excel_datetime called with dt_val={dt_val} (type: {type(dt_val)}), date_val_for_time={date_val_for_time}, is_just_date={is_just_date}")

    if pd.isna(dt_val) or dt_val is None: return None
    if isinstance(dt_val, datetime): return dt_val.date() if is_just_date else dt_val
    if isinstance(dt_val, date) and not isinstance(dt_val, datetime): return dt_val 
    
    if isinstance(dt_val, time) and not is_just_date: 
        if date_val_for_time and isinstance(date_val_for_time, date):
            result = datetime.combine(date_val_for_time, dt_val)
            # if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 1 (input is time), returning {result} (type {type(result)})")
            return result
        return None 
    
    if isinstance(dt_val, (int, float)): 
        try:
            dt_obj_from_openpyxl = openpyxl_from_excel(dt_val)
            
            # KIỂM TRA NẾU openpyxl_from_excel TRẢ VỀ TIME OBJECT
            if isinstance(dt_obj_from_openpyxl, time):
                if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2.A (openpyxl returned time: {dt_obj_from_openpyxl})")
                if not is_just_date and date_val_for_time and isinstance(date_val_for_time, date):
                    result = datetime.combine(date_val_for_time, dt_obj_from_openpyxl)
                    if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2.A.1 (combined with ref_date), returning {result} (type {type(result)})")
                    return result
                elif is_just_date: # Muốn date từ time object là không hợp lý
                    if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2.A.2 (is_just_date for time obj), returning None")
                    return None
                else: # Không có ref_date để combine
                    if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2.A.3 (no ref_date for time obj), returning None (hoặc có thể là time_obj nếu nghiệp vụ cho phép)")
                    return None # Hoặc trả về dt_obj_from_openpyxl nếu muốn chấp nhận time object đơn lẻ
            
            # Nếu dt_obj_from_openpyxl là datetime (ví dụ: từ số serial date lớn)
            elif isinstance(dt_obj_from_openpyxl, datetime):
                result = dt_obj_from_openpyxl.date() if is_just_date else dt_obj_from_openpyxl
                if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2.B (openpyxl returned datetime), returning {result} (type {type(result)})")
                return result
            else: # Trường hợp khác không mong muốn từ openpyxl_from_excel
                if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2.C (unexpected type from openpyxl: {type(dt_obj_from_openpyxl)}), returning None")
                return None

        except (ValueError, TypeError): # Lỗi khi openpyxl_from_excel không parse được (ví dụ số quá lớn/nhỏ)
            if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 2 (except after openpyxl_from_excel failed - điều này không nên xảy ra với 0.75)")
            # Khối này có thể không cần thiết nữa nếu logic trên đã xử lý time fraction
            # Tuy nhiên, để an toàn, giữ lại để bắt các lỗi khác của openpyxl_from_excel
            # if 0 <= dt_val < 1: # Điều kiện này có thể đã được xử lý bởi isinstance(dt_obj_from_openpyxl, time)
            #    ...
            return None 
            
    if isinstance(dt_val, str):
        dt_s = dt_val.strip()
        if not dt_s or dt_s.lower() in ["from", "to", "hours", "error", "remark", "-", "/"]: return None 
        result = _parse_datetime_from_string_logic(dt_s, date_val_for_time, is_just_date, context)
        # if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 3 (string), returning {result} (type {type(result)})")
        return result

    # if dt_val == 0.75: print(f"DEBUG_FUNC: Branch 4 (default end), returning None")
    return None

def parse_time_duration(time_val) -> float: # Parses duration into hours (float)
    # ... (Giữ nguyên code hàm này từ extract_updated.py)
    if isinstance(time_val, timedelta): return timedelta_to_hours(time_val)
    if pd.isna(time_val) or time_val is None: return 0.0

    if isinstance(time_val, (datetime, time)): 
        val = time_val.hour + time_val.minute / 60 + time_val.second / 3600
        return round(val, 2)

    if isinstance(time_val, str):
        try: 
            parts = list(map(int, re.split('[:.]', time_val)))
            val = 0.0
            if len(parts) == 2: 
                val = parts[0] + parts[1] / 60
            elif len(parts) == 3: 
                val = parts[0] + parts[1] / 60 + parts[2] / 3600
            return round(val, 2)
        except ValueError:
            logging.warning(f"parse_time_duration: Cannot parse duration string: '{time_val}'")
            return 0.0
            
    if isinstance(time_val, (int, float)): 
        val = time_val * 24 if 0 <= time_val < 1 else float(time_val)
        return round(val, 2)

    logging.warning(f"parse_time_duration: Unsupported type '{type(time_val)}' for duration value '{time_val}'")
    return 0.0